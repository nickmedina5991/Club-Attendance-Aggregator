"""
Club Attendance Aggregator
==========================
Reads one or more CSV files exported from EngageSUNY / CampusGroups
and produces (or updates) an Excel workbook tracking each member's
total attendance count, sorted highest to lowest, with a checkmark column
for every meeting date scanned.
 
Expected CSV structure (UBLinked export):
    Line 1: "Event Attendance By Event"
    Line 2: blank
    Line 3: Event name
    Line 4: Start Date, <date>
    Line 5: End Date,   <date>
    Line 6: First Name, Last Name, Campus Email, ... 
    Line 7+: data rows
 
Output columns:
    First Name | Last Name | Campus Email | Attendance Count | <date1> | <date2> | ...
 
Usage:
    python attendance_aggregator.py GBM_02_20_2026.csv GBM_02_27_2026.csv ...
    python attendance_aggregator.py *.csv
    python attendance_aggregator.py          # processes every *.csv in current directory
"""

import sys, glob, os
from datetime import datetime
from io import StringIO
 
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
OUTPUT_FILE = "attendance_summary.xlsx"
 
# CSV parsing
def parse_csv(path: str) -> tuple:
    """
    Returns (dataframe, event_name, meeting_date).
    dataframe has columns: First Name, Last Name, Campus Email.
    Returns (None, "", "") on failure.
    """
    with open(path, encoding="utf-8-sig") as f:
        lines = f.readlines()
 
    header_idx = None
    for i, line in enumerate(lines):
        if "first name" in line.lower():
            header_idx = i
            break
 
    if header_idx is None:
        print(f"    !!! Skipping '{path}' - could not find 'First Name' header row.")
        return None, "", ""
 
    event_name = lines[2].strip().strip(",").strip('"') if len(lines) > 2 else ""
 
    meeting_date = ""
    for line in lines[:header_idx]:
        if line.lower().startswith("start date"):
            parts = line.split(",", 1)
            if len(parts) == 2:
                meeting_date = parts[1].strip().strip('"')
            break
 
    if not meeting_date:
        try:
            mtime = os.path.getmtime(path)
            meeting_date = datetime.fromtimestamp(mtime).strftime("%m/%d/%Y")
        except Exception:
            meeting_date = datetime.today().strftime("%m/%d/%Y")
 
    data_text = "".join(lines[header_idx:])
    df = pd.read_csv(StringIO(data_text), dtype=str).fillna("")
    df.columns = [c.strip() for c in df.columns]
 
    needed = {"First Name", "Last Name", "Campus Email"}
    missing = needed - set(df.columns)
    if missing:
        print(f"    !!! Skipping '{path}' - missing columns: {missing}")
        return None, "", ""
 
    df = df[["First Name", "Last Name", "Campus Email"]].copy()
    df = df.map(str.strip)
    return df, event_name, meeting_date
 
 
# Summary persistence
def load_existing(path: str) -> tuple:
    """
    Load existing summary Excel.
    Returns:
        records   - dict keyed by normalised email
        dates     - ordered list of date strings already in the sheet
    Each record has: First Name, Last Name, Campus Email, Attendance Count,
                     and one key per date (e.g. "1/23/2026") -> "checkmark" or ""
    """
    if not os.path.exists(path):
        return {}, []
 
    try:
        df = pd.read_excel(path, sheet_name="Summary", dtype=str)
    except Exception:
        return {}, []
 
    # Fixed columns; everything after is a date column
    fixed = {"First Name", "Last Name", "Campus Email", "Attendance Count"}
    date_cols = [c for c in df.columns if c not in fixed and not c.startswith("Unnamed")]
 
    records = {}
    for _, row in df.iterrows():
        email = str(row.get("Campus Email", "")).strip().lower()
        if not email or email == "nan":
            continue
        try:
            count = int(float(str(row.get("Attendance Count", 0))))
        except ValueError:
            count = 0
 
        rec = {
            "First Name":       str(row.get("First Name", "")).strip(),
            "Last Name":        str(row.get("Last Name", "")).strip(),
            "Campus Email":     str(row.get("Campus Email", "")).strip(),
            "Attendance Count": count,
        }
        for d in date_cols:
            raw = row.get(d, "")
            # pandas reads empty Excel cells as NaN; treat anything that isn't "checkmark" as blank
            rec[d] = "✓" if str(raw).strip() == "✓" else ""
 
        records[email] = rec
 
    return records, date_cols
 
 
def load_import_log(path: str) -> set:
    """
    Read the Import Log sheet and return a set of already-processed filenames.
    Returns an empty set if the file or sheet does not exist.
    """
    if not os.path.exists(path):
        return set()
    try:
        df = pd.read_excel(path, sheet_name="Import Log", dtype=str)
        return set(df["File"].dropna().str.strip().tolist())
    except Exception:
        return set()
 
 
# Processing & merging
def process_csvs(csv_files: list) -> tuple:
    """
    Returns:
        incremental  - dict keyed by normalised email
        new_dates    - ordered list of new date strings encountered (deduped)
        log_entries  - list of (filename, event_name, meeting_date)
    """
    incremental = {}
    new_dates   = []
    log_entries = []
 
    for path in csv_files:
        print(f"  Reading: {path}")
        df, event_name, meeting_date = parse_csv(path)
        if df is None:
            continue
 
        if meeting_date not in new_dates:
            new_dates.append(meeting_date)
 
        log_entries.append((os.path.basename(path), event_name, meeting_date))
 
        for _, row in df.iterrows():
            email = row["Campus Email"].strip().lower()
            if not email:
                continue
 
            if email not in incremental:
                incremental[email] = {
                    "First Name":       row["First Name"],
                    "Last Name":        row["Last Name"],
                    "Campus Email":     row["Campus Email"],
                    "Attendance Count": 0,
                }
 
            incremental[email]["Attendance Count"]  += 1
            incremental[email][meeting_date]         = "✓"
            if row["First Name"]: incremental[email]["First Name"] = row["First Name"]
            if row["Last Name"]:  incremental[email]["Last Name"]  = row["Last Name"]
 
        print(f"    -> {len(df)} attendees  |  '{event_name}'  |  {meeting_date}")
 
    return incremental, new_dates, log_entries
 
 
def merge(existing: dict, existing_dates: list, incremental: dict, new_dates: list) -> tuple:
    """
    Merge incremental data into existing records.
    Returns (merged_records, all_dates_ordered) with dates sorted oldest to newest.
    """
    all_dates = list(existing_dates)
    for d in new_dates:
        if d not in all_dates:
            all_dates.append(d)
    all_dates.sort(key=lambda d: datetime.strptime(d, "%m/%d/%Y"))
 
    merged = {}
    for key, rec in existing.items():
        merged[key] = dict(rec)
        for d in all_dates:
            merged[key].setdefault(d, "")
 
    for key, inc_rec in incremental.items():
        if key in merged:
            merged[key]["Attendance Count"] += inc_rec["Attendance Count"]
            if inc_rec["First Name"]: merged[key]["First Name"] = inc_rec["First Name"]
            if inc_rec["Last Name"]:  merged[key]["Last Name"]  = inc_rec["Last Name"]
            for d in new_dates:
                if inc_rec.get(d) == "✓":
                    merged[key][d] = "✓"
        else:
            new_rec = dict(inc_rec)
            for d in all_dates:
                new_rec.setdefault(d, "")
            merged[key] = new_rec
 
    for key in merged:
        for d in all_dates:
            merged[key].setdefault(d, "")
 
    return merged, all_dates
 
 
# Styling
HDR_FILL   = PatternFill("solid", start_color="1F4E79")
DATE_FILL  = PatternFill("solid", start_color="2E75B6")
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_FILL   = PatternFill("solid", start_color="D6E4F0")
NORM_FILL  = PatternFill("solid", start_color="FFFFFF")
TOT_FILL   = PatternFill("solid", start_color="D9D9D9")
CHECK_FONT = Font(name="Arial", size=11, color="375623", bold=True)
BODY_FONT  = Font(name="Arial", size=10)
BOLD_FONT  = Font(name="Arial", size=10, bold=True)
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",   vertical="center")
BORDER     = Border(
    left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin",  color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"),
)
 
def _style(cell, font, fill, align):
    cell.font = font; cell.fill = fill; cell.alignment = align; cell.border = BORDER
 
def style_header_row(ws, ncols, date_start_col):
    for c in range(1, ncols + 1):
        fill = DATE_FILL if c >= date_start_col else HDR_FILL
        _style(ws.cell(1, c), HDR_FONT, fill, CENTER)
 
def style_body_row(ws, ri, ncols, count_col, date_start_col):
    fill = ALT_FILL if ri % 2 == 0 else NORM_FILL
    for c in range(1, ncols + 1):
        cell = ws.cell(ri, c)
        if c >= date_start_col and cell.value == "✓":
            _style(cell, CHECK_FONT, fill, CENTER)
        elif c == count_col:
            _style(cell, BODY_FONT, fill, CENTER)
        else:
            _style(cell, BODY_FONT, fill, LEFT)
 
 
# Excel writer
def write_excel(records: dict, all_dates: list, path: str, log_entries: list) -> None:
    wb = Workbook()
 
    ws = wb.active
    ws.title = "Summary"
 
    fixed_headers = ["First Name", "Last Name", "Campus Email", "Attendance Count"]
    headers       = fixed_headers + all_dates
    count_col     = 4
    date_start    = len(fixed_headers) + 1
 
    ws.append(headers)
    style_header_row(ws, len(headers), count_col)
    ws.row_dimensions[1].height = 22
 
    sorted_recs = sorted(
        records.values(),
        key=lambda r: (-r["Attendance Count"], r["Last Name"].lower(), r["First Name"].lower())
    )
 
    for ri, rec in enumerate(sorted_recs, start=2):
        ws.cell(ri, 1, rec["First Name"])
        ws.cell(ri, 2, rec["Last Name"])
        ws.cell(ri, 3, rec["Campus Email"])
        ws.cell(ri, 4, rec["Attendance Count"])
        for di, d in enumerate(all_dates, start=date_start):
            ws.cell(ri, di, rec.get(d, ""))
        style_body_row(ws, ri, len(headers), count_col, date_start)
 
    tr = len(sorted_recs) + 2
    ws.cell(tr, 1, "TOTAL MEMBERS")
    ws.cell(tr, count_col, f"=COUNTA(C2:C{tr-1})")
    for di, _ in enumerate(all_dates, start=date_start):
        col_letter = get_column_letter(di)
        ws.cell(tr, di, f'=COUNTIF({col_letter}2:{col_letter}{tr-1},"✓")')
 
    for c in range(1, len(headers) + 1):
        cell = ws.cell(tr, c)
        cell.fill = TOT_FILL; cell.border = BORDER; cell.font = BOLD_FONT
        cell.alignment = CENTER if c >= count_col else LEFT
 
    for col, w in zip("ABCD", [22, 22, 22, 22]):
        ws.column_dimensions[col].width = w
    for di in range(date_start, date_start + len(all_dates)):
        ws.column_dimensions[get_column_letter(di)].width = 13
 
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
 
    log = wb.create_sheet("Import Log")
    log_hdrs = ["File", "Event Name", "Meeting Date", "Imported At"]
    log.append(log_hdrs)
    style_header_row(log, len(log_hdrs), date_start_col=99)
 
    # Load existing log rows so they are preserved across runs
    existing_log_rows = []
    if os.path.exists(path):
        try:
            prev = pd.read_excel(path, sheet_name="Import Log", dtype=str).fillna("")
            for _, row in prev.iterrows():
                existing_log_rows.append((
                    str(row.get("File", "")).strip(),
                    str(row.get("Event Name", "")).strip(),
                    str(row.get("Meeting Date", "")).strip(),
                    str(row.get("Imported At", "")).strip(),
                ))
        except Exception:
            pass
 
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    all_log_rows = existing_log_rows + [(fname, event, date, now) for fname, event, date in log_entries]
 
    for ri, (fname, event, date, imported_at) in enumerate(all_log_rows, start=2):
        log.cell(ri, 1, fname); log.cell(ri, 2, event)
        log.cell(ri, 3, date);  log.cell(ri, 4, imported_at)
        fill = ALT_FILL if ri % 2 == 0 else NORM_FILL
        for c in range(1, 5):
            _style(log.cell(ri, c), BODY_FONT, fill, LEFT)
 
    for col, w in zip("ABCD", [28, 40, 16, 18]):
        log.column_dimensions[col].width = w
 
    wb.save(path)
 
 
# Main
def main():
    if len(sys.argv) > 1:
        csv_files = []
        for arg in sys.argv[1:]:
            csv_files.extend(glob.glob(arg))
    else:
        csv_files = glob.glob("*.csv")
 
    csv_files = sorted(set(csv_files))
 
    if not csv_files:
        print("No CSV files found. Pass paths as arguments or run in a folder with *.csv files.")
        sys.exit(1)
 
    print(f"\nFound {len(csv_files)} CSV file(s).\n")
 
    print("Loading existing summary...")
    existing, existing_dates = load_existing(OUTPUT_FILE)
    already_imported = load_import_log(OUTPUT_FILE)
    print(f"  {len(existing)} existing member(s), {len(existing_dates)} date(s) already tracked.")
 
    if already_imported:
        before = len(csv_files)
        csv_files = [f for f in csv_files if os.path.basename(f) not in already_imported]
        skipped = before - len(csv_files)
        if skipped:
            print(f"  Skipping {skipped} already-imported file(s). Check Import Log for details.")
 
    if not csv_files:
        print("\n  Nothing new to import. Exiting.\n")
        return
 
    print()
    print("Processing CSVs...")
    incremental, new_dates, log_entries = process_csvs(csv_files)
    print(f"\n  {len(incremental)} unique member(s) in new files.")
 
    print("\nMerging...")
    merged, all_dates = merge(existing, existing_dates, incremental, new_dates)
    print(f"  {len(merged)} total member(s) | {len(all_dates)} total date(s) tracked.")
 
    print(f"\nWriting '{OUTPUT_FILE}'...")
    write_excel(merged, all_dates, OUTPUT_FILE, log_entries)
    print(f"  Done -> {os.path.abspath(OUTPUT_FILE)}\n")
 
 
if __name__ == "__main__":
    main()